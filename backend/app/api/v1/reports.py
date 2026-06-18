from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from io import BytesIO
from datetime import date

from app.core.database import get_db
from app.core.security import get_current_user
from app.models.user import User
from app.services.report_service import ReportService

router = APIRouter()
report_service = ReportService()


@router.get("/daily")
def daily_report(
    project_id: int,
    report_date: str = Query(...),
    format: str = "json",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        rdate = date.fromisoformat(report_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD.")

    data = report_service.generate_daily_report(project_id, rdate, db)

    if format == "pdf":
        pdf = generate_pdf(data, "Daily Site Report")
        return StreamingResponse(BytesIO(pdf), media_type="application/pdf",
                                 headers={"Content-Disposition": f"attachment; filename=daily_report_{report_date}.pdf"})
    elif format == "excel":
        excel = generate_excel(data, "Daily Site Report")
        return StreamingResponse(BytesIO(excel), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": f"attachment; filename=daily_report_{report_date}.xlsx"})

    return data


@router.get("/weekly")
def weekly_report(
    project_id: int,
    end_date: str = Query(...),
    format: str = "json",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        edate = date.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD.")

    data = report_service.generate_weekly_report(project_id, edate, db)
    return filter_format(data, "Weekly Report", format)


@router.get("/monthly")
def monthly_report(
    project_id: int,
    year: int = Query(...),
    month: int = Query(...),
    format: str = "json",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not (1 <= month <= 12):
        raise HTTPException(status_code=400, detail="Invalid month. Must be 1-12.")

    data = report_service.generate_monthly_report(project_id, year, month, db)
    return filter_format(data, "Monthly Report", format)


@router.get("/labour")
def labour_report(
    project_id: int,
    date_from: str = Query(...),
    date_to: str = Query(...),
    format: str = "json",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        dfrom = date.fromisoformat(date_from)
        dto = date.fromisoformat(date_to)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD.")

    data = report_service.generate_labour_report(project_id, dfrom, dto, db)
    return filter_format(data, "Labour Report", format)


@router.get("/material")
def material_report(
    project_id: int,
    date_from: str = Query(...),
    date_to: str = Query(...),
    format: str = "json",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        dfrom = date.fromisoformat(date_from)
        dto = date.fromisoformat(date_to)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD.")

    data = report_service.generate_material_report(project_id, dfrom, dto, db)
    return filter_format(data, "Material Report", format)


@router.get("/cost")
def cost_report(
    project_id: int,
    date_from: str = Query(...),
    date_to: str = Query(...),
    format: str = "json",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        dfrom = date.fromisoformat(date_from)
        dto = date.fromisoformat(date_to)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD.")

    data = report_service.generate_cost_report(project_id, dfrom, dto, db)
    return filter_format(data, "Cost Report", format)


def filter_format(data: dict, title: str, format: str):
    if format == "pdf":
        pdf = generate_pdf(data, title)
        return StreamingResponse(BytesIO(pdf), media_type="application/pdf",
                                 headers={"Content-Disposition": f"attachment; filename=report.pdf"})
    elif format == "excel":
        excel = generate_excel(data, title)
        return StreamingResponse(BytesIO(excel),
                                 media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": f"attachment; filename=report.xlsx"})
    return data


def generate_pdf(data: dict, title: str) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from io import BytesIO

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(f"<b>{title}</b>", styles["Title"]))
    story.append(Spacer(1, 12))
    story.append(Paragraph(f"<b>Project:</b> {data.get('project_name', 'N/A')}", styles["Normal"]))
    story.append(Paragraph(f"<b>Date/Period:</b> {data.get('date', data.get('period', 'N/A'))}", styles["Normal"]))
    story.append(Spacer(1, 12))

    for key, value in data.items():
        if key in ("report_type", "project_name", "date", "period"):
            continue

        story.append(Paragraph(f"<b>{key.replace('_', ' ').title()}</b>", styles["Heading2"]))
        
        if isinstance(value, list):
            if not value:
                story.append(Paragraph("<i>No records found</i>", styles["Normal"]))
                story.append(Spacer(1, 12))
            elif isinstance(value[0], dict):
                headers = list(value[0].keys())
                table_data = [[h.replace("_", " ").title() for h in headers]]
                for item in value:
                    table_data.append([str(item.get(h, "")) for h in headers])
                t = Table(table_data)
                t.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.5, colors.grey), ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey)]))
                story.append(t)
                story.append(Spacer(1, 12))
            else:
                story.append(Paragraph(f"{value}", styles["Normal"]))
                story.append(Spacer(1, 12))
        elif isinstance(value, dict):
            for k, v in value.items():
                if isinstance(v, list):
                    story.append(Spacer(1, 6))
                    story.append(Paragraph(f"<b>{k.replace('_', ' ').title()}</b>", styles["Heading3"]))
                    if not v:
                        story.append(Paragraph("<i>None</i>", styles["Normal"]))
                    elif isinstance(v[0], dict):
                        headers = list(v[0].keys())
                        table_data = [[h.replace("_", " ").title() for h in headers]]
                        for item in v:
                            table_data.append([str(item.get(h, "")) for h in headers])
                        t = Table(table_data)
                        t.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.5, colors.grey), ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey)]))
                        story.append(t)
                    else:
                        story.append(Paragraph(f"{v}", styles["Normal"]))
                else:
                    story.append(Paragraph(f"<b>{k.replace('_', ' ').title()}:</b> {v}", styles["Normal"]))
            story.append(Spacer(1, 12))
        else:
            story.append(Paragraph(f"{value}", styles["Normal"]))
            story.append(Spacer(1, 12))

    doc.build(story)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


def generate_excel(data: dict, title: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

    ws.cell(1, 1, title).font = Font(bold=True, size=14)
    ws.cell(2, 1, f"Period: {data.get('date', data.get('period', 'N/A'))}")

    row = 4
    for key, value in data.items():
        if key in ("report_type", "project_name", "date", "period"):
            continue

        if isinstance(value, list):
            ws.cell(row, 1, key.replace("_", " ").title()).font = Font(bold=True)
            row += 1
            if not value:
                ws.cell(row, 1, "No records found").font = Font(italic=True)
                row += 2
            elif isinstance(value[0], dict):
                headers = list(value[0].keys())
                for col, h in enumerate(headers, 1):
                    cell = ws.cell(row, col, h.replace("_", " ").title())
                    cell.font = header_font
                    cell.fill = header_fill
                row += 1
                for item in value:
                    for col, h in enumerate(headers, 1):
                        ws.cell(row, col, str(item.get(h, "")))
                    row += 1
                row += 1
            else:
                ws.cell(row, 1, str(value))
                row += 2
        elif isinstance(value, dict):
            # Summary Block
            ws.cell(row, 1, key.replace("_", " ").title()).font = Font(bold=True)
            row += 1
            for k, v in value.items():
                if isinstance(v, list):
                    ws.cell(row, 1, k.replace("_", " ").title()).font = Font(bold=True)
                    row += 1
                    if not v:
                        ws.cell(row, 1, "None").font = Font(italic=True)
                        row += 1
                    elif isinstance(v[0], dict):
                        headers = list(v[0].keys())
                        for col, h in enumerate(headers, 1):
                            cell = ws.cell(row, col, h.replace("_", " ").title())
                            cell.font = header_font
                            cell.fill = header_fill
                        row += 1
                        for item in v:
                            for col, h in enumerate(headers, 1):
                                ws.cell(row, col, str(item.get(h, "")))
                            row += 1
                    else:
                        ws.cell(row, 1, str(v))
                        row += 1
                else:
                    ws.cell(row, 1, k.replace("_", " ").title())
                    ws.cell(row, 2, str(v))
                    row += 1
            row += 1
        else:
            # Primitive
            ws.cell(row, 1, key.replace("_", " ").title()).font = Font(bold=True)
            ws.cell(row, 2, str(value))
            row += 1

    buffer = BytesIO()
    wb.save(buffer)
    excel_bytes = buffer.getvalue()
    buffer.close()
    return excel_bytes
